from collections import OrderedDict
from os import path

import numpy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.utils import get_column_letter

from casium import adb

def trimmean(arr, percent):
    if len(arr) == 0:
        return 0

    sorted_arr = sorted(arr)
    n = len(sorted_arr)
    k = int(round(n * percent))
    return numpy.mean(sorted_arr[k:n - k])

def trimsd(arr, percent):
    if len(arr) == 0:
        return 0

    sorted_arr = sorted(arr)
    n = len(sorted_arr)
    k = int(round(n * percent))
    trim_arr = sorted_arr[k:n - k]

    mean = trimmean(arr, percent)
    deviation_arr = [(v - mean) * (v - mean) for v in trim_arr]
    return numpy.sqrt(numpy.mean(deviation_arr))

def cmd_start_response_info(udid):
    adb.shell('atrace gfx input --async_start', udid)

def cmd_stop_response_info(udid):
    return adb.shell('atrace gfx input --async_stop', udid)[1]

class ResponseData(object):
    def __init__(self, row_data, package):
        self.row_data = row_data
        self.package = package
        self.input_datas = []
        self.surface_flinger_datas = []

    def parse_row_data(self):
        rows = self.row_data.split('\n')
        for row in rows:
            input_pos = row.find('InputDispatcher')
            if input_pos != -1:
                time = self._get_row_timestamp(row)
                self.input_datas.append(time)

            surface_pos = row.find('surfaceflinger')
            package_pos = row.find(self.package)
            if surface_pos != -1 and package_pos != -1:
                time = self._get_row_timestamp(row)
                self.surface_flinger_datas.append(time)
        # calculate response time and surface flinger time
        if self.input_datas and self.surface_flinger_datas:
            self.response_time = self.surface_flinger_datas[0] - self.input_datas[-1]
            self.surface_flinger_time = self.surface_flinger_datas[-1] - self.surface_flinger_datas[0]
        else:
            self.response_time = -1
            self.surface_flinger_time = -1

    def _get_row_timestamp(self, row):
        time_str_end = row.find(': tracing_mark_write')
        time_str = row[time_str_end - 13:time_str_end]
        time = float(time_str) * 1000  # time unit : ms
        return time

class ResponseInfo(object):
    def __init__(self, package_name):
        self._package_name = package_name
        self._datas = []

    def trace_begin(self, udid):
        cmd_start_response_info(udid)

    def trace_end(self, udid):
        data = ResponseData(cmd_stop_response_info(udid), self._package_name)
        self._datas += [data]
        data.parse_row_data()

    def dump_to_metrics(self, metrics):
        MetricsWriter(metrics, self._datas, self).dump()

    def write_to_excel(self, name):
        ExcelWriter(name, self._datas).dump()

class MetricsWriter(object):
    def __init__(self, metrics, datas, responseinfo):
        self._metrics = metrics
        self._datas = datas
        self._responseinfo = responseinfo

    def dump(self):
        values = OrderedDict()

        avg_response = trimmean([d.response_time for d in self._datas], 0.1)
        sd_response = trimsd([d.response_time for d in self._datas], 0.1)
        values['Response'] = (avg_response, sd_response)

        avg_surface_flinger = trimmean([d.surface_flinger_time for d in self._datas], 0.1)
        sd_surface_flinger = trimsd([d.surface_flinger_time for d in self._datas], 0.1)
        values['Surface Flinger'] = (avg_surface_flinger, sd_surface_flinger)

        avg_total = trimmean([(d.response_time + d.surface_flinger_time) for d in self._datas], 0.1)
        sd_total = trimsd([(d.response_time + d.surface_flinger_time) for d in self._datas], 0.1)
        values['Total'] = (avg_total, sd_total)

        self._metrics.add('response', values, self._responseinfo)

class ExcelWriter(object):
    sum_row = {'next': 2}

    def __init__(self, name, datas):
        self._file_name = 'responseinfo.xlsx'
        self._name = name
        self._datas = datas

    def dump(self):
        col_titles = OrderedDict([('response_time', 'Response Time(ms)'),
                                  ('surface_flinger_time', 'Surface Flinger Time(ms)'),
                                  ('total_time', 'Total Time(ms)')])

        if path.isfile(self._file_name):
            wb = load_workbook(self._file_name)
        else:
            wb = Workbook()
        ws = wb.create_sheet(title=self._name)

        ft_bold = Font(bold=True)

        # Set titles
        for col, name in enumerate(col_titles.values()):
            col_letter = get_column_letter(col + 1)
            ws['%s%s' % (col_letter, 1)] = name
            ws['%s%s' % (col_letter, 1)].font = ft_bold
            ws.column_dimensions[col_letter].width = len(name) + 4

        # Set values
        for row, data in enumerate(self._datas):
            ws['%s%s' % (get_column_letter(1), row + 2)] = data.response_time
            ws['%s%s' % (get_column_letter(2), row + 2)] = data.surface_flinger_time
            ws['%s%s' % (get_column_letter(3), row + 2)] = (data.response_time + data.surface_flinger_time)

        # Set express info
        sum_ws = wb.active
        sum_ws.title = 'summary'

        # Express titles
        express_titles = ['Test Case',
                          'Response Time(ms)(avg)',
                          'Response Time(ms)(sd)',
                          'Surface Flinger Time(ms)(avg)',
                          'Surface Flinger Time(ms)(sd)',
                          'Total Time(ms)(avg)',
                          'Total Time(ms)(sd)']

        for col, name in enumerate(express_titles):
            col_letter = get_column_letter(col + 1)
            sum_ws['%s%s' % (col_letter, 1)] = name
            sum_ws['%s%s' % (col_letter, 1)].font = ft_bold
            sum_ws.column_dimensions[col_letter].width = len(name) + 4

        row = self.sum_row['next']




        avg_surface_flinger = trimmean([d.surface_flinger_time for d in self._datas], 0.1)
        sd_surface_flinger = trimsd([d.surface_flinger_time for d in self._datas], 0.1)

        avg_total = trimmean([(d.response_time + d.surface_flinger_time) for d in self._datas], 0.1)
        sd_total = trimsd([(d.response_time + d.surface_flinger_time) for d in self._datas], 0.1)

        sum_ws['A%s' % row] = self._name
        avg_response = trimmean([d.response_time for d in self._datas], 0.1)
        sd_response = trimsd([d.response_time for d in self._datas], 0.1)
        sum_ws['B%s' % row] = avg_response
        sum_ws['B%s' % row].number_format = FORMAT_NUMBER_00
        sum_ws['C%s' % row] = sd_response
        sum_ws['C%s' % row].number_format = FORMAT_NUMBER_00
        avg_surface_flinger = trimmean([d.surface_flinger_time for d in self._datas], 0.1)
        sd_surface_flinger = trimsd([d.surface_flinger_time for d in self._datas], 0.1)
        sum_ws['D%s' % row] = avg_surface_flinger
        sum_ws['D%s' % row].number_format = FORMAT_NUMBER_00
        sum_ws['E%s' % row] = sd_surface_flinger
        sum_ws['E%s' % row].number_format = FORMAT_NUMBER_00
        avg_total = trimmean([(d.response_time + d.surface_flinger_time) for d in self._datas], 0.1)
        sd_total = trimsd([(d.response_time + d.surface_flinger_time) for d in self._datas], 0.1)
        sum_ws['F%s' % row] = avg_total
        sum_ws['F%s' % row].number_format = FORMAT_NUMBER_00
        sum_ws['G%s' % row] = sd_total
        sum_ws['G%s' % row].number_format = FORMAT_NUMBER_00

        self.sum_row['next'] += 1

        wb.save(self._file_name)
