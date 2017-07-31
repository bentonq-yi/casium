import threading
import time
from collections import OrderedDict
from os import path

import numpy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00
from openpyxl.utils import get_column_letter

from casium import adb


def trimmean(arr, percent):
    if len(arr) == 0:
        return 0

    sorted_arr = sorted(arr)
    n = len(sorted_arr)
    k = int(round(n * percent))
    return numpy.mean(sorted_arr[k:n - k])


def trimsd(arr, percent):
    if len(arr) == 0:
        return 0

    sorted_arr = sorted(arr)
    n = len(sorted_arr)
    k = int(round(n * percent))
    trim_arr = sorted_arr[k:n - k]

    mean = trimmean(arr, percent)
    deviation_arr = [(v - mean) * (v - mean) for v in trim_arr]
    return numpy.sqrt(numpy.mean(deviation_arr))


def trimmax(arr, percent):
    if len(arr) == 0:
        return 0

    sorted_arr = sorted(arr)
    n = len(sorted_arr)
    k = int(round(n * percent))
    return round(sorted_arr[n - k - 1])


def cmd_reset_gfx_info(package_name, udid):
    adb.shell('dumpsys gfxinfo %s reset' % package_name, udid)


def cmd_dump_gfx_info(package_name, udid):
    return adb.shell('dumpsys gfxinfo %s' % package_name, udid)[1]


class GfxData(object):
    def __init__(self, row_data):
        self.row_data = row_data
        self.pid = int(self._extract_value('Graphics info for pid', '['))
        self.pkg = self._extract_value('[', ']')
        self.stats_since_ns = int(self._extract_value('Stats since:', 'ns'))
        self.total_frames = int(self._extract_value('Total frames rendered:'))
        self.janky_frames = int(self._extract_value('Janky frames:', '('))
        self.percentile_50th = int(self._extract_value('50th percentile:', 'ms'))
        self.percentile_90th = int(self._extract_value('90th percentile:', 'ms'))
        self.percentile_95th = int(self._extract_value('95th percentile:', 'ms'))
        self.percentile_99th = int(self._extract_value('99th percentile:', 'ms'))
        self.missed_vsync = int(self._extract_value('Number Missed Vsync:'))
        self.high_input_latency = int(self._extract_value('Number High input latency:'))
        self.slow_ui_thread = int(self._extract_value('Number Slow UI thread:'))
        self.slow_bitmap_uploads = int(self._extract_value('Number Slow bitmap uploads:'))
        self.slow_issue_draw_commands = int(self._extract_value('Number Slow issue draw commands:'))
        self.histogram = self._extract_histogram('HISTOGRAM:')

        if self.total_frames != 0:
            self.janky_rate = float(self.janky_frames) / self.total_frames
        else:
            self.janky_rate = 0.0

        self.frame_ms = []
        for key, value in self.histogram.iteritems():
            for i in range(value):
                self.frame_ms += [key]

        self.mean_50th = trimmean(self.frame_ms, 0.25)
        self.mean_90th = trimmean(self.frame_ms, 0.05)
        self.mean_95th = trimmean(self.frame_ms, 0.025)
        self.mean_99th = trimmean(self.frame_ms, 0.005)

        self.sd_50th = trimsd(self.frame_ms, 0.25)
        self.sd_90th = trimsd(self.frame_ms, 0.05)
        self.sd_95th = trimsd(self.frame_ms, 0.025)
        self.sd_99th = trimsd(self.frame_ms, 0.005)

    def _extract_value(self, start_key, end_key='\n'):
        s = self.row_data.index(start_key)
        e = self.row_data.index(end_key, s)
        return self.row_data[s + len(start_key): e].strip()

    def _extract_histogram(self, start_key, end_key='\n'):
        v = self._extract_value(start_key, end_key)
        return dict((int(k.partition('ms')[0]), int(v)) for k, v in
                    (it.split('=') for it in v.split(' ')))


class GfxInfo(object):
    def __init__(self, package_name):
        self._package_name = package_name
        self._datas = []

    def trace_begin(self, udid):
        cmd_reset_gfx_info(self._package_name, udid)

    def trace_end(self, udid):
        self._datas += [GfxData(cmd_dump_gfx_info(self._package_name, udid))]

    def dump_to_metrics(self, metrics):
        MetricsWriter(metrics, self._datas, self).dump()

    def write_to_excel(self, name, device):
        ExcelWriter(name, self._datas, device).dump()


class GfxData2(object):
    def __init__(self):
        self.frame_ms = []
        self.ui_draw_ms = []
        self.prepare_ms = []
        self.process_ms = []
        self.execute_ms = []
        self.total_frames = 0
        self.janky_frames = 0
        self.janky_rate = 0
        self.mean_50th = 0
        self.mean_90th = 0
        self.mean_95th = 0
        self.mean_99th = 0
        self.sd_50th = 0
        self.sd_90th = 0
        self.sd_95th = 0
        self.sd_99th = 0
        self.percentile_50th = 0
        self.percentile_90th = 0
        self.percentile_95th = 0
        self.percentile_99th = 0

    def add_row_data(self, row_data):
        profile_data = self._extract_value(row_data, 'Profile data in ms:', 'View hierarchy')
        frames_ms = profile_data.split('\n')[2:]
        wash_frames_ms = [frame_ms for frame_ms in frames_ms if len(frame_ms.split('\t')) == 5]
        ui_draw_ms = [float(frame_ms.split('\t')[1]) for frame_ms in wash_frames_ms]
        prepare_ms = [float(frame_ms.split('\t')[2]) for frame_ms in wash_frames_ms]
        process_ms = [float(frame_ms.split('\t')[3]) for frame_ms in wash_frames_ms]
        execute_ms = [float(frame_ms.split('\t')[4]) for frame_ms in wash_frames_ms]
        frame_ms = []
        for i in range(len(ui_draw_ms)):
            frame_ms += [ui_draw_ms[i] + prepare_ms[i] + process_ms[i] + execute_ms[i]]

        self.frame_ms += frame_ms
        self.ui_draw_ms += ui_draw_ms
        self.prepare_ms += prepare_ms
        self.process_ms += process_ms
        self.execute_ms += execute_ms

    def parse_row_data(self):
        self.total_frames = len(self.frame_ms)
        self.janky_frames = len([ms for ms in self.frame_ms if ms > 16.666])
        if self.total_frames != 0:
            self.janky_rate = float(self.janky_frames) / self.total_frames
        self.mean_50th = trimmean(self.frame_ms, 0.25)
        self.mean_90th = trimmean(self.frame_ms, 0.05)
        self.mean_95th = trimmean(self.frame_ms, 0.025)
        self.mean_99th = trimmean(self.frame_ms, 0.005)
        self.sd_50th = trimsd(self.frame_ms, 0.25)
        self.sd_90th = trimsd(self.frame_ms, 0.05)
        self.sd_95th = trimsd(self.frame_ms, 0.025)
        self.sd_99th = trimsd(self.frame_ms, 0.005)
        self.percentile_50th = trimmax(self.frame_ms, 0.5)
        self.percentile_90th = trimmax(self.frame_ms, 0.1)
        self.percentile_95th = trimmax(self.frame_ms, 0.05)
        self.percentile_99th = trimmax(self.frame_ms, 0.01)

    def _extract_value(self, row_data, start_key, end_key='\n'):
        s = row_data.index(start_key)
        e = row_data.index(end_key, s)
        return row_data[s + len(start_key): e].strip()

    def _extract_histogram(self, start_key, end_key='\n'):
        v = self._extract_value(start_key, end_key)
        return dict((int(k.partition('ms')[0]), int(v)) for k, v in
                    (it.split('=') for it in v.split(' ')))


class GfxThread(threading.Thread):
    def __init__(self, gfxinfo2):
        threading.Thread.__init__(self)
        self._gfxinfo2 = gfxinfo2

    def run(self):
        self._gfxinfo2.run_gfx_trace()


class GfxInfo2(object):
    def __init__(self, package_name):
        self._continue_trace = False
        self._package_name = package_name
        self._datas = []
        self._gfx_data = None
        self._trace_thread = None
        self._udid = None

    def trace_begin(self, udid):
        cmd_reset_gfx_info(self._package_name, udid)
        self._continue_trace = True
        self._gfx_data = GfxData2()
        self._trace_thread = GfxThread(self)
        self._trace_thread.start()
        self._udid = udid

    def trace_end(self, udid):
        self._continue_trace = False
        self._trace_thread.join()
        self._gfx_data.parse_row_data()
        self._datas += [self._gfx_data]

    def run_gfx_trace(self):
        while self._continue_trace:
            self._gfx_data.add_row_data(cmd_dump_gfx_info(self._package_name, self._udid))
            time.sleep(1)

    def dump_to_metrics(self, metrics):
        MetricsWriter(metrics, self._datas, self).dump()

    def write_to_excel(self, name, device):
        ExcelWriter(name, self._datas, device).dump()


class MetricsWriter(object):
    def __init__(self, metrics, datas, gfxinfo):
        self._metrics = metrics
        self._datas = datas
        self._gfxinfo = gfxinfo

    def dump(self):
        values = OrderedDict()

        avg_janky_rate = trimmean([d.janky_rate for d in self._datas], 0.1)
        values['Janky Rate'] = (avg_janky_rate, 0)

        avg_mean_90th = trimmean([d.mean_90th for d in self._datas], 0.1)
        avg_sd_90th = trimmean([d.sd_90th for d in self._datas], 0.1)
        values['90th Percentile'] = (avg_mean_90th, avg_sd_90th)

        avg_mean_95th = trimmean([d.mean_95th for d in self._datas], 0.1)
        avg_sd_95th = trimmean([d.sd_95th for d in self._datas], 0.1)
        values['95th Percentile'] = (avg_mean_95th, avg_sd_95th)

        avg_mean_99th = trimmean([d.mean_99th for d in self._datas], 0.1)
        avg_sd_99th = trimmean([d.sd_99th for d in self._datas], 0.1)
        values['99th Percentile'] = (avg_mean_99th, avg_sd_99th)

        self._metrics.add('gfx', values, self._gfxinfo)


class ExcelWriter(object):
    sum_row = {'next': 11}

    def __init__(self, name, datas, device):
        self._file_name = "gfx.xlsx"
        # if not name.endswith('.xlsx'):
        #     self._file_name = '%s.xlsx' % name
        self._name = name
        self._datas = datas
        self._device = device

    def dump(self):
        col_titles = OrderedDict([('total_frames', 'Total Frames'),
                                  ('janky_frames', 'Janky Frames'),
                                  ('janky_rate', 'Janky Rate'),
                                  ('mean_90th', '90th Percentile(avg)'),
                                  ('sd_90th', '90th Percentile(sd)'),
                                  ('percentile_90th', '90th Percentile(max)'),
                                  ('mean_95th', '95th Percentile(avg)'),
                                  ('sd_95th', '95th Percentile(sd)'),
                                  ('percentile_95th', '95th Percentile(max)'),
                                  ('mean_99th', '99th Percentile(avg)'),
                                  ('sd_99th', '99th Percentile(sd)'),
                                  ('percentile_99th', '99th Percentile(max)')])

        if path.isfile(self._file_name):
            wb = load_workbook(self._file_name)
        else:
            wb = Workbook()
        ws = wb.create_sheet(title=self._name)

        ft_bold = Font(bold=True)

        # Set titles
        for col, name in enumerate(col_titles.values()):
            col_letter = get_column_letter(col + 1)
            ws['%s%s' % (col_letter, 1)] = name
            ws['%s%s' % (col_letter, 1)].font = ft_bold
            ws.column_dimensions[col_letter].width = len(name) + 4

        # Set values
        for row, data in enumerate(self._datas):
            for col, prop in enumerate(col_titles.keys()):
                col_letter = get_column_letter(col + 1)
                ws['%s%s' % (col_letter, row + 2)] = getattr(data, prop, None)
                if col == 2:  # janky_rate
                    ws['%s%s' % (col_letter, row + 2)].number_format = FORMAT_PERCENTAGE_00
                elif col > 2:
                    ws['%s%s' % (col_letter, row + 2)].number_format = FORMAT_NUMBER_00

        sum_ws = wb.active
        sum_ws.title = 'summary'

        # Summary titles
        summary_titles = ['Test Case',
                          'Janky Rate',
                          '90th Percentile(avg)',
                          '90th Percentile(sd)',
                          '95th Percentile(avg)',
                          '95th Percentile(sd)',
                          '99th Percentile(avg)',
                          '99th Percentile(sd)']

        for col, name in enumerate(summary_titles):
            col_letter = get_column_letter(col + 1)
            sum_ws['%s%s' % (col_letter, 10)] = name
            sum_ws['%s%s' % (col_letter, 10)].font = ft_bold
            sum_ws.column_dimensions[col_letter].width = len(name) + 4

        row = self.sum_row['next']

        sum_ws['A%s' % row] = self._name
        avg_janky_rate = trimmean([d.janky_rate for d in self._datas], 0.1)
        sum_ws['B%s' % row] = avg_janky_rate
        sum_ws['B%s' % row].number_format = FORMAT_PERCENTAGE_00

        avg_mean_90th = trimmean([d.mean_90th for d in self._datas], 0.1)
        avg_sd_90th = trimmean([d.sd_90th for d in self._datas], 0.1)
        sum_ws['C%s' % row] = avg_mean_90th
        sum_ws['C%s' % row].number_format = FORMAT_NUMBER_00
        sum_ws['D%s' % row] = avg_sd_90th
        sum_ws['D%s' % row].number_format = FORMAT_NUMBER_00

        avg_mean_95th = trimmean([d.mean_95th for d in self._datas], 0.1)
        avg_sd_95th = trimmean([d.sd_95th for d in self._datas], 0.1)
        sum_ws['E%s' % row] = avg_mean_95th
        sum_ws['E%s' % row].number_format = FORMAT_NUMBER_00
        sum_ws['F%s' % row] = avg_sd_95th
        sum_ws['F%s' % row].number_format = FORMAT_NUMBER_00

        avg_mean_99th = trimmean([d.mean_99th for d in self._datas], 0.1)
        avg_sd_99th = trimmean([d.sd_99th for d in self._datas], 0.1)
        sum_ws['G%s' % row] = avg_mean_99th
        sum_ws['G%s' % row].number_format = FORMAT_NUMBER_00
        sum_ws['H%s' % row] = avg_sd_99th
        sum_ws['H%s' % row].number_format = FORMAT_NUMBER_00

        self.sum_row['next'] += 1

        # Save target info
        sum_ws['A1'] = 'Device'
        sum_ws['B1'] = self._device.product_model
        sum_ws['A2'] = 'Package Name'
        sum_ws['B2'] = self._device.current_package_name
        sum_ws['A3'] = 'Activity Name'
        sum_ws['B3'] = self._device.current_activity_name
        sum_ws['A4'] = 'Test Time'
        sum_ws['B4'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

        wb.save(self._file_name)
